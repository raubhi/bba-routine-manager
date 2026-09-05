import streamlit as st
import pandas as pd
import json
import re
import copy
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from ortools.sat.python import cp_model

st.set_page_config(page_title="BBA Routine Optimizer - Complete Matrix", layout="wide")

# --- COPYRIGHT HEADER ---
st.markdown("""
<div style='text-align: center; padding: 15px; background-color: #1F4E78; color: white; border-radius: 8px; margin-bottom: 20px;'>
    <h1 style='margin: 0; font-size: 28px;'>BBA Routine Management System</h1>
    <p style='margin: 5px 0 0 0; font-size: 14px; color: #E9ECEF;'>
        © Copyrighted to <b>Rezwan Aubhi</b> | <a href='mailto:r.aubhi@gmail.com' style='color: #FFD700; text-decoration: none;'>r.aubhi@gmail.com</a>
    </p>
</div>
""", unsafe_allow_html=True)

# --- INITIALIZE PERSISTENT SESSION STATES ---
if 'fac_rules' not in st.session_state: st.session_state.fac_rules = []
if 'batch_rules' not in st.session_state: st.session_state.batch_rules = {}
if 'sun_mon_facs' not in st.session_state: st.session_state.sun_mon_facs = []
if 'fixed_rooms' not in st.session_state: st.session_state.fixed_rooms = {}
if 'preview_data' not in st.session_state: st.session_state.preview_data = None
if 'routine_history' not in st.session_state: st.session_state.routine_history = []
if 'current_file_id' not in st.session_state: st.session_state.current_file_id = None
if 'base_fac_df' not in st.session_state: st.session_state.base_fac_df = None
if 'latest_fac_df_dict' not in st.session_state: st.session_state.latest_fac_df_dict = None
if 'pending_fix_data' not in st.session_state: st.session_state.pending_fix_data = None
if 'pending_violations' not in st.session_state: st.session_state.pending_violations = []

# --- CONFIGURATION MANAGER ---
st.subheader("💾 Configuration Manager (Save/Load Rules)")
col_conf1, col_conf2 = st.columns(2)
with col_conf1:
    config_dict = {
        "fac_rules": st.session_state.fac_rules,
        "batch_rules": st.session_state.batch_rules,
        "sun_mon_facs": st.session_state.sun_mon_facs,
        "fixed_rooms": st.session_state.fixed_rooms,
        "fac_matrix": st.session_state.latest_fac_df_dict
    }
    st.download_button("📥 Download Saved Rules (.json)", data=json.dumps(config_dict), file_name="bba_routine_rules.json", mime="application/json")
with col_conf2:
    uploaded_config = st.file_uploader("📤 Upload Saved Rules (.json)", type=["json"])
    if uploaded_config:
        loaded_conf = json.load(uploaded_config)
        st.session_state.fac_rules = loaded_conf.get("fac_rules", [])
        st.session_state.batch_rules = loaded_conf.get("batch_rules", {})
        st.session_state.sun_mon_facs = loaded_conf.get("sun_mon_facs", [])
        st.session_state.fixed_rooms = loaded_conf.get("fixed_rooms", {})
        if loaded_conf.get("fac_matrix"):
            st.session_state.base_fac_df = pd.DataFrame(loaded_conf["fac_matrix"])
            if "fac_editor_widget" in st.session_state: del st.session_state["fac_editor_widget"]
        st.success("Rules loaded successfully!")

st.divider()

# --- 1. DATA UPLOAD & STATEFUL PARSING ---
st.subheader("1. Data Upload & Master Parsing")
course_file = st.file_uploader("Upload Fall 2026 Course Offering Sheet", type=["xlsx"])

if course_file:
    file_id = f"{course_file.name}_{course_file.size}"
    is_new_file = False
    if st.session_state.current_file_id != file_id:
        st.session_state.current_file_id = file_id
        is_new_file = True

    xls = pd.ExcelFile(course_file)
    target_sheet = next((s for s in xls.sheet_names if 'bba' in s.lower() or 'offer' in s.lower()), xls.sheet_names[0])
    df_raw = pd.read_excel(xls, sheet_name=target_sheet, header=None)
    
    parsed_rows = []
    current_batch = "20th Batch"
    last_code = None
    last_title = "Unknown Course"
    
    for idx, row in df_raw.iterrows():
        row_vals = [str(v).strip() for v in row.values if pd.notna(v) and str(v).strip() != '']
        if not row_vals: continue
        row_str = " ".join(row_vals).lower()
        
        if "batch" in row_str and len(row_vals) < 4:
            for v in row_vals:
                if "batch" in v.lower():
                    m = re.search(r'(\d+(?:st|nd|rd|th)?)', v, flags=re.IGNORECASE)
                    if m:
                        current_batch = f"{m.group(1).title()} Batch"
                    else:
                        current_batch = v.split(',')[0].replace(":", "").replace("Sections", "").strip().title()
            continue
            
        code_val = next((v for v in row_vals if any(c.isdigit() for c in v) and '-' in v and len(v) >= 5), None)
        
        if code_val:
            last_code = code_val
            code_idx = row_vals.index(code_val)
            last_title = row_vals[code_idx + 1] if code_idx + 1 < len(row_vals) else "Unknown Course"
            data_cells = row_vals[code_idx+2:]
        else:
            if last_code and len(row_vals) >= 1 and not any(kw in row_str for kw in ['major', 'minor', 'total', 'credit', 'batch']):
                data_cells = row_vals
            else:
                continue
                
        if not data_cells: data_cells = ["TBA", "A"]
        last_val = str(data_cells[-1]).upper()
        sections = ["A"]
        
        if not any(w in last_val for w in ['SEM', 'YEAR', 'CRED', 'TOTAL', 'BATCH', 'TEACHER']):
            sec_text = last_val.replace('SECTIONS', '').replace('SECTION', '').replace('SEC', '').replace(':', '')
            sec_clean = sec_text.replace(',', ' ').replace(';', ' ').replace('&', ' ').replace('AND', ' ').replace('-', ' ')
            tokens = [t.strip() for t in sec_clean.split() if t.strip()]
            if all(len(t) <= 3 and t.isalnum() for t in tokens) and len(tokens) > 0:
                sections = tokens
                search_space = data_cells[:-1]
            else: search_space = data_cells
        else: search_space = data_cells
                
        teacher = "TBA"
        for cell in reversed(search_space):
            cl = cell.lower()
            if any(w in cl for w in ['semester', 'year', 'cred', 'th', 'st', 'nd', 'rd', 'batch']): continue 
            if len(cell) > 2:
                teacher = cell
                break
                
        for sec in sections:
            parsed_rows.append({
                "Batch_Core": current_batch,
                "Batch": f"{current_batch} (Sec {sec})",
                "Section": sec,
                "Code": last_code,
                "Title": last_title,
                "Faculty": teacher,
                "Credits": 3
            })

    df_tasks = pd.DataFrame(parsed_rows)

    if is_new_file or st.session_state.base_fac_df is None:
        faculty_sheet = next((s for s in xls.sheet_names if 'faculty' in s.lower()), None)
        faculty_data = []
        if faculty_sheet:
            df_fac = pd.read_excel(xls, sheet_name=faculty_sheet)
            is_adjunct = False
            for _, row in df_fac.iterrows():
                first_col = str(row.iloc[0]).strip().lower()
                if first_col == 'contractual':
                    is_adjunct = True
                    continue
                name = str(row.iloc[1]).strip()
                if name.lower() not in ['nan', 'name', '']:
                    faculty_data.append({"Faculty Name": name, "Type": 'Adjunct' if is_adjunct else 'Full-Time'})
        
        if not faculty_data:
            extracted = df_tasks['Faculty'].dropna().unique()
            faculty_data = [{"Faculty Name": str(f).strip(), "Type": "Full-Time"} for f in extracted if str(f).strip() != '']
            
        if not any(f['Faculty Name'] == 'TBA' for f in faculty_data):
            faculty_data.append({"Faculty Name": "TBA", "Type": "Adjunct"})
            
        fac_df = pd.DataFrame(faculty_data)
        for day in ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]:
            fac_df[f"{day} Max"] = 3
            
        st.session_state.base_fac_df = fac_df
        if "fac_editor_widget" in st.session_state: del st.session_state["fac_editor_widget"]

    days_ordered = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
    slots_ordered = ["9:30-11:00", "11:00-12:30", "1:30-3:00", "3:00-4:30"]

    st.subheader("2. Persistent Faculty Roster & Daily Limits Matrix")
    col_config = {"Type": st.column_config.SelectboxColumn("Type", options=["Full-Time", "Adjunct"], required=True)}
    for day in days_ordered:
        col_config[f"{day} Max"] = st.column_config.NumberColumn(f"{day} Max", min_value=0, max_value=6, step=1)
    
    edited_faculty = st.data_editor(st.session_state.base_fac_df, num_rows="dynamic", column_config=col_config, use_container_width=True, key="fac_editor_widget")
    st.session_state.latest_fac_df_dict = edited_faculty.to_dict('records')
    available_faculties = edited_faculty[edited_faculty['Faculty Name'] != 'TBA']['Faculty Name'].tolist()
    
    st.divider()

    # --- 2. ADVANCED RULES (TABBED ENVIRONMENT) ---
    st.subheader("3. Advanced Scheduling Constraints")
    tab_locks, tab_batches, tab_sunmon, tab_rooms = st.tabs(["Faculty Locks", "Batch Limits & Blackouts", "Force Sun/Mon Classes", "Room Mappings"])
    
    with tab_locks:
        col_t1, col_t2 = st.columns([1, 2])
        with col_t1:
            selected_fac = st.selectbox("Select Faculty:", available_faculties, key="l_fac")
            rule_day = st.selectbox("Select Day:", days_ordered, key="l_day")
            rule_slot = st.selectbox("Select Time Slot:", slots_ordered, key="l_slot")
            if st.button("➕ Lock Faculty to Slot"):
                st.session_state.fac_rules.append({"Faculty": selected_fac, "Day": rule_day, "Slot": rule_slot})
                st.rerun()
        with col_t2:
            if st.session_state.fac_rules:
                st.table(pd.DataFrame(st.session_state.fac_rules))
                if st.button("Clear All Faculty Locks"):
                    st.session_state.fac_rules = []
                    st.rerun()
            
    with tab_batches:
        col_b1, col_b2 = st.columns([1, 2])
        with col_b1:
            target_batches = st.multiselect("Select Batch(es):", df_tasks['Batch'].unique().tolist())
            batch_max_days = st.number_input(f"Max Active Days:", min_value=1, max_value=6, value=3)
            blocked_batch_days = st.multiselect(f"Strictly Block Days:", days_ordered)
            if st.button("💾 Save Batch Configuration"):
                for b in target_batches: st.session_state.batch_rules[b] = {"Max Days": batch_max_days, "Blocked Days": blocked_batch_days}
                st.rerun()
        with col_b2:
            if st.session_state.batch_rules:
                st.json(st.session_state.batch_rules)
                if st.button("Clear All Batch Rules"):
                    st.session_state.batch_rules = {}
                    st.rerun()

    with tab_sunmon:
        st.write("**Force Classes to Sunday & Monday**")
        col_sm1, col_sm2 = st.columns([1, 2])
        with col_sm1:
            sm_fac = st.selectbox("Select Faculty:", available_faculties, key="sm_fac_sel")
            if st.button("➕ Force Sun/Mon Load"):
                if sm_fac not in st.session_state.sun_mon_facs:
                    st.session_state.sun_mon_facs.append(sm_fac)
                    st.rerun()
        with col_sm2:
            if st.session_state.sun_mon_facs:
                st.write("**Currently Forced Faculties:**", st.session_state.sun_mon_facs)
                if st.button("Clear Sun/Mon Locks"):
                    st.session_state.sun_mon_facs = []
                    st.rerun()

    with tab_rooms:
        st.write("**Fix Room Numbers Batch-Wise (Optional)**")
        col_r1, col_r2 = st.columns([1, 2])
        with col_r1:
            r_batch = st.selectbox("Select Batch to Lock:", df_tasks['Batch_Core'].unique().tolist())
            r_num = st.text_input("Assign Room (e.g., Room 101):")
            if st.button("💾 Lock Room to Batch"):
                st.session_state.fixed_rooms[r_batch] = r_num
                st.rerun()
        with col_r2:
            if st.session_state.fixed_rooms:
                st.json(st.session_state.fixed_rooms)
                if st.button("Clear Room Locks"):
                    st.session_state.fixed_rooms = {}
                    st.rerun()

    st.divider()

    # --- 3. GLOBAL ENGINE SETTINGS ---
    st.subheader("4. Global Engine Settings")
    col_g1, col_g2 = st.columns(2)
    with col_g1:
        include_majors = st.toggle("Include Major Courses", value=False)
        compact_schedule = st.toggle("Enforce Compact Scheduling", value=False)
        replace_tba = st.toggle("Convert TBA slots to an assigned Faculty member?", value=False)
        tba_replacement = st.selectbox("TBA Replacement Faculty:", available_faculties) if replace_tba else None
    with col_g2:
        total_rooms = st.number_input("Total Available Rooms per Time Slot:", min_value=1, max_value=50, value=15, step=1)

    st.divider()
    
    # --- AUTO-FIX DASHBOARD (Shows only if fixes are pending) ---
    if st.session_state.pending_fix_data is not None:
        st.error("❌ **Strict constraints produced a mathematical contradiction.**")
        st.info("💡 **Diagnostic AI Found a Fix!** To generate a clash-free routine, the system suggests safely overriding the following bottlenecks:")
        
        for v in st.session_state.pending_violations:
            st.warning(v)
            
        col_acc, col_rej = st.columns(2)
        with col_acc:
            if st.button("✅ Accept Fixes & Apply Routine", type="primary", use_container_width=True):
                st.session_state.preview_data = st.session_state.pending_fix_data
                st.session_state.routine_history = [copy.deepcopy(st.session_state.pending_fix_data)]
                st.session_state.pending_fix_data = None
                st.session_state.pending_violations = []
                st.rerun()
        with col_rej:
            if st.button("❌ Reject & Adjust Manually", use_container_width=True):
                st.session_state.pending_fix_data = None
                st.session_state.pending_violations = []
                st.rerun()
        st.divider()

    # --- 4. SOLVER EXECUTION ---
    if st.button("🚀 Generate Full Multi-Batch Master Routine", type="primary", use_container_width=True):
        
        master_tasks = []
        for _, row in df_tasks.iterrows():
            if not include_majors and "major" in row['Title'].lower(): continue
            teacher = row['Faculty']
            if teacher.lower() == 'tba' and replace_tba and tba_replacement: teacher = tba_replacement
            master_tasks.append({
                "Batch_Core": row['Batch_Core'], "Batch": row['Batch'], "Code": row['Code'],
                "Title": row['Title'], "Faculty": teacher, "Section": row['Section']
            })

        active_facs = list(set(t["Faculty"] for t in master_tasks if t["Faculty"] != "TBA"))
            
        with st.spinner("Compiling Math Engine and Optimizing..."):
            
            # --- STRICT SOLVER PASS ---
            model = cp_model.CpModel()
            x = {}
            for i in range(len(master_tasks)):
                for d in days_ordered:
                    for s in slots_ordered: x[(i, d, s)] = model.NewBoolVar(f"x_{i}_{d}_{s}")
                        
            for i in range(len(master_tasks)):
                model.AddExactlyOne(x[(i, d, s)] for d in days_ordered for s in slots_ordered)
                
            for d in days_ordered:
                for s in slots_ordered:
                    model.Add(sum(x[(i, d, s)] for i in range(len(master_tasks))) <= total_rooms)
                    batch_sec_map = {}
                    for i, t in enumerate(master_tasks): batch_sec_map.setdefault(t['Batch'], []).append(i)
                    for indices in batch_sec_map.values(): model.AddAtMostOne(x[(i, d, s)] for i in indices)
                        
                    fac_map = {}
                    for i, t in enumerate(master_tasks):
                        if t["Faculty"] != "TBA": fac_map.setdefault(t["Faculty"], []).append(i)
                    for indices in fac_map.values(): model.AddAtMostOne(x[(i, d, s)] for i in indices)

            for _, f_row in edited_faculty.iterrows():
                fac_name = f_row["Faculty Name"]
                if fac_name == "TBA": continue
                f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == fac_name]
                for d in days_ordered: model.Add(sum(x[(i, d, s)] for i in f_indices for s in slots_ordered) <= int(f_row[f"{d} Max"]))

            for rule in st.session_state.fac_rules:
                f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == rule["Faculty"]]
                if f_indices: model.Add(sum(x[(i, rule["Day"], rule["Slot"])] for i in f_indices) == 1)

            for sm_fac in st.session_state.sun_mon_facs:
                f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == sm_fac]
                if f_indices:
                    tot = len(f_indices)
                    s_tgt = min(3, tot)
                    m_tgt = min(3, max(0, tot - s_tgt))
                    model.Add(sum(x[(i, "Sunday", s)] for i in f_indices for s in slots_ordered) == s_tgt)
                    model.Add(sum(x[(i, "Monday", s)] for i in f_indices for s in slots_ordered) == m_tgt)

            for b_key, rules in st.session_state.batch_rules.items():
                b_indices = [i for i, t in enumerate(master_tasks) if t["Batch"] == b_key]
                if not b_indices: continue
                for d in rules["Blocked Days"]:
                    for i in b_indices:
                        for s in slots_ordered: model.Add(x[(i, d, s)] == 0)
                b_active_vars = []
                for d in days_ordered:
                    day_active = model.NewBoolVar(f"b_act_{b_key}_{d}")
                    model.AddMaxEquality(day_active, [x[(i, d, s)] for i in b_indices for s in slots_ordered])
                    b_active_vars.append(day_active)
                model.Add(sum(b_active_vars) <= rules["Max Days"])

            if compact_schedule:
                for f in active_facs:
                    f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == f]
                    for d in days_ordered:
                        s1 = sum(x[(i, d, slots_ordered[0])] for i in f_indices)
                        s2 = sum(x[(i, d, slots_ordered[1])] for i in f_indices)
                        s3 = sum(x[(i, d, slots_ordered[2])] for i in f_indices)
                        s4 = sum(x[(i, d, slots_ordered[3])] for i in f_indices)
                        model.Add(s1 + s4 <= 1 + s2 + s3)

            solver = cp_model.CpSolver()
            solver.parameters.max_time_in_seconds = 10.0
            status = solver.Solve(model)
            
            if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                # STRICT PASS SUCCESS
                scheduled_output = []
                for i, t in enumerate(master_tasks):
                    for d in days_ordered:
                        for s in slots_ordered:
                            if solver.Value(x[(i, d, s)]) == 1:
                                assigned_room = st.session_state.fixed_rooms.get(t['Batch_Core'], f"Room {i % total_rooms + 1}")
                                scheduled_output.append({"ID": f"Task_{i}", "Batch": t['Batch'], "Batch_Core": t['Batch_Core'], "Section": t['Section'], "Day": d, "Time Slot": s, "Course Code": t['Code'], "Course Title": t['Title'], "Faculty": t['Faculty'], "Room": assigned_room})
                st.session_state.preview_data = scheduled_output
                st.session_state.routine_history = [copy.deepcopy(scheduled_output)]
                st.success("Routine Generated Automatically! Zero conflicts found.")
                st.rerun()

            else:
                # --- AUTO-FIX RELAXATION SOLVER PASS ---
                f_model = cp_model.CpModel()
                fx = {}
                for i in range(len(master_tasks)):
                    for d in days_ordered:
                        for s in slots_ordered: fx[(i, d, s)] = f_model.NewBoolVar(f"fx_{i}_{d}_{s}")
                            
                for i in range(len(master_tasks)): f_model.AddExactlyOne(fx[(i, d, s)] for d in days_ordered for s in slots_ordered)
                    
                penalties = []
                fac_slacks = {}
                batch_blk_slacks = {}
                batch_day_slacks = {}
                room_slacks = {}

                for d in days_ordered:
                    for s in slots_ordered:
                        r_slk = f_model.NewIntVar(0, 50, f"r_slk_{d}_{s}")
                        f_model.Add(sum(fx[(i, d, s)] for i in range(len(master_tasks))) <= total_rooms + r_slk)
                        penalties.append(r_slk * 100)
                        room_slacks[(d,s)] = r_slk
                        
                        batch_sec_map = {}
                        for i, t in enumerate(master_tasks): batch_sec_map.setdefault(t['Batch'], []).append(i)
                        for indices in batch_sec_map.values(): f_model.AddAtMostOne(fx[(i, d, s)] for i in indices)
                            
                        fac_map = {}
                        for i, t in enumerate(master_tasks):
                            if t["Faculty"] != "TBA": fac_map.setdefault(t["Faculty"], []).append(i)
                        for indices in fac_map.values(): f_model.AddAtMostOne(fx[(i, d, s)] for i in indices)

                for _, f_row in edited_faculty.iterrows():
                    fac_name = f_row["Faculty Name"]
                    if fac_name == "TBA": continue
                    f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == fac_name]
                    for d in days_ordered:
                        f_slk = f_model.NewIntVar(0, 10, f"fslk_{fac_name}_{d}")
                        f_model.Add(sum(fx[(i, d, s)] for i in f_indices for s in slots_ordered) <= int(f_row[f"{d} Max"]) + f_slk)
                        penalties.append(f_slk * 50)
                        fac_slacks[(fac_name, d)] = f_slk

                for rule in st.session_state.fac_rules:
                    f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == rule["Faculty"]]
                    if f_indices: f_model.Add(sum(fx[(i, rule["Day"], rule["Slot"])] for i in f_indices) == 1)

                for sm_fac in st.session_state.sun_mon_facs:
                    f_indices = [i for i, t in enumerate(master_tasks) if t["Faculty"] == sm_fac]
                    if f_indices:
                        tot = len(f_indices)
                        s_tgt = min(3, tot)
                        m_tgt = min(3, max(0, tot - s_tgt))
                        sun_slk = f_model.NewIntVar(0, 3, f"sun_slk_{sm_fac}")
                        mon_slk = f_model.NewIntVar(0, 3, f"mon_slk_{sm_fac}")
                        f_model.Add(sum(fx[(i, "Sunday", s)] for i in f_indices for s in slots_ordered) >= s_tgt - sun_slk)
                        f_model.Add(sum(fx[(i, "Monday", s)] for i in f_indices for s in slots_ordered) >= m_tgt - mon_slk)
                        penalties.extend([sun_slk * 30, mon_slk * 30])

                for b_key, rules in st.session_state.batch_rules.items():
                    b_indices = [i for i, t in enumerate(master_tasks) if t["Batch"] == b_key]
                    if not b_indices: continue
                    for d in rules["Blocked Days"]:
                        b_blk_slk = f_model.NewIntVar(0, 10, f"bblk_{b_key}_{d}")
                        f_model.Add(sum(fx[(i, d, s)] for i in b_indices for s in slots_ordered) <= b_blk_slk)
                        penalties.append(b_blk_slk * 80)
                        batch_blk_slacks[(b_key, d)] = b_blk_slk
                        
                    b_active_vars = []
                    for d in days_ordered:
                        day_active = f_model.NewBoolVar(f"fb_act_{b_key}_{d}")
                        f_model.AddMaxEquality(day_active, [fx[(i, d, s)] for i in b_indices for s in slots_ordered])
                        b_active_vars.append(day_active)
                        
                    b_day_slk = f_model.NewIntVar(0, 6, f"bdayslk_{b_key}")
                    f_model.Add(sum(b_active_vars) <= rules["Max Days"] + b_day_slk)
                    penalties.append(b_day_slk * 60)
                    batch_day_slacks[b_key] = b_day_slk

                f_model.Minimize(sum(penalties))
                f_solver = cp_model.CpSolver()
                f_solver.parameters.max_time_in_seconds = 15.0
                f_status = f_solver.Solve(f_model)

                if f_status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                    violations = []
                    for (d, s), var in room_slacks.items():
                        if f_solver.Value(var) > 0: violations.append(f"🏢 Needed **{f_solver.Value(var)} extra room(s)** on {d} at {s}.")
                    for (fac, d), var in fac_slacks.items():
                        if f_solver.Value(var) > 0: violations.append(f"👨‍🏫 **{fac}** was scheduled for **{f_solver.Value(var)} extra class(es)** on {d} beyond their Daily Limit Matrix.")
                    for (b, d), var in batch_blk_slacks.items():
                        if f_solver.Value(var) > 0: violations.append(f"🎓 **{b}** had to be scheduled on **{d}** (which was set as a Blocked Day).")
                    for b, var in batch_day_slacks.items():
                        if f_solver.Value(var) > 0: violations.append(f"🎓 **{b}** required **{f_solver.Value(var)} extra active day(s)** beyond its limit to finish the syllabus.")

                    scheduled_output = []
                    for i, t in enumerate(master_tasks):
                        for d in days_ordered:
                            for s in slots_ordered:
                                if f_solver.Value(fx[(i, d, s)]) == 1:
                                    assigned_room = st.session_state.fixed_rooms.get(t['Batch_Core'], f"Room {i % total_rooms + 1}")
                                    scheduled_output.append({"ID": f"Task_{i}", "Batch": t['Batch'], "Batch_Core": t['Batch_Core'], "Section": t['Section'], "Day": d, "Time Slot": s, "Course Code": t['Code'], "Course Title": t['Title'], "Faculty": t['Faculty'], "Room": assigned_room})
                    
                    st.session_state.pending_violations = violations
                    st.session_state.pending_fix_data = scheduled_output
                    st.rerun()
                else:
                    st.error("❌ Hard Clash Detected. One of your batches or faculties has more than 24 total classes assigned, which makes generation physically impossible. Check your Excel file.")

    # --- 5. INTERACTIVE DASHBOARDS ---
    if st.session_state.preview_data is not None:
        days_ordered = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
        slots_ordered = ["9:30-11:00", "11:00-12:30", "1:30-3:00", "3:00-4:30"]
        
        tab_grid, tab_facs, tab_swap = st.tabs(["📊 Master Grid View", "👨‍🏫 Faculty Profiles", "🔄 Manual Swap & Undo Engine"])
        
        with tab_grid:
            batches = sorted(list(set(row["Batch"] for row in st.session_state.preview_data)))
            grid = {d: {s: {b: "" for b in batches} for s in slots_ordered} for d in days_ordered}
            for row in st.session_state.preview_data:
                b, d, s = row["Batch"], row["Day"], row["Time Slot"]
                grid[d][s][b] = f"<b>{row['Course Code']}</b><br>{row['Course Title']}<br>{row['Faculty']}<br><i>{row['Room']}</i>"

            html = "<div style='overflow-x:auto;'><table style='width:100%; border-collapse: collapse; text-align: center; font-size: 13px; font-family: sans-serif;'>"
            html += "<tr><th style='border: 1px solid #ccc; background-color: #1F4E78; color: white; padding: 10px;'>Day</th>"
            html += "<th style='border: 1px solid #ccc; background-color: #1F4E78; color: white; padding: 10px;'>Time Slot</th>"
            for b in batches: html += f"<th style='border: 1px solid #ccc; background-color: #1F4E78; color: white; padding: 10px;'>{b}</th>"
            html += "</tr>"
            for d in days_ordered:
                for i, s in enumerate(slots_ordered):
                    html += "<tr>"
                    if i == 0: html += f"<td rowspan='4' style='border: 1px solid #ccc; font-weight: bold; background-color: #f8f9fa; vertical-align: middle;'>{d}</td>"
                    html += f"<td style='border: 1px solid #ccc; background-color: #e9ecef; padding: 8px; white-space: nowrap;'>{s}</td>"
                    for b in batches:
                        val = grid[d][s][b]
                        bg = "#ffffff" if val else "#fafafa"
                        html += f"<td style='border: 1px solid #ccc; padding: 10px; background-color: {bg}; min-width: 150px;'>{val}</td>"
                    html += "</tr>"
            html += "</table></div>"
            st.markdown(html, unsafe_allow_html=True)
            
            # EXCEL EXPORT
            wb = Workbook()
            ws = wb.active
            ws.title = "Master Routine"
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(batches)+2)
            title_cell = ws.cell(row=1, column=1, value="UNIVERSITY OF SCHOLARS - BBA PROGRAM ROUTINE (FALL 2026)")
            title_cell.font = Font(name='Arial', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=2, column=1, value="Day").fill = header_fill
            ws.cell(row=2, column=1).font = header_font
            ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=2, column=1).border = thin_border
            
            ws.cell(row=2, column=2, value="Time Slot").fill = header_fill
            ws.cell(row=2, column=2).font = header_font
            ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=2, column=2).border = thin_border
            
            for idx, b in enumerate(batches, 3):
                c = ws.cell(row=2, column=idx, value=b)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = thin_border
                ws.column_dimensions[c.column_letter].width = 25
                
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            
            row_idx = 3
            for d in days_ordered:
                start_row = row_idx
                for s in slots_ordered:
                    sc = ws.cell(row=row_idx, column=2, value=s)
                    sc.alignment = Alignment(horizontal='center', vertical='center')
                    sc.border = thin_border
                    for c_idx, b in enumerate(batches, 3):
                        val = grid[d][s][b].replace("<b>", "").replace("</b>", "").replace("<br>", "\n").replace("<i>", "").replace("</i>", "")
                        cell = ws.cell(row=row_idx, column=c_idx, value=val)
                        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                        cell.border = thin_border
                    row_idx += 1
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row_idx-1, end_column=1)
                dc = ws.cell(row=start_row, column=1, value=d)
                dc.alignment = Alignment(horizontal='center', vertical='center')
                dc.font = Font(bold=True)
                for r in range(start_row, row_idx): ws.cell(row=r, column=1).border = thin_border
                
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            st.download_button("📥 Download Official Master Routine (.xlsx)", data=output, file_name="BBA_Fall_2026_Master_Routine.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

        with tab_facs:
            st.write("### Faculty Workload & Isolated Routines")
            active_facs = sorted(list(set(row["Faculty"] for row in st.session_state.preview_data)))
            
            prof_data = []
            for f in active_facs:
                f_courses = [r for r in st.session_state.preview_data if r["Faculty"] == f]
                b_list = list(set([r["Batch_Core"] for r in f_courses]))
                s_list = list(set([r["Section"] for r in f_courses]))
                credits = len(f_courses) * 3
                desig = edited_faculty[edited_faculty['Faculty Name'] == f]['Type'].values[0] if f in edited_faculty['Faculty Name'].values else "Unknown"
                prof_data.append({"Faculty Name": f, "Designation": desig, "Allocated Batches": ", ".join(b_list), "Sections": ", ".join(s_list), "Total Credits": credits})
            
            st.dataframe(pd.DataFrame(prof_data), use_container_width=True)
            
            selected_facs = st.multiselect("Select Faculty to compare visual routines:", active_facs)
            if selected_facs:
                fac_grid = {d: {s: {f: "" for f in selected_facs} for s in slots_ordered} for d in days_ordered}
                for row in st.session_state.preview_data:
                    fac = row["Faculty"]
                    if fac in selected_facs:
                        fac_grid[row["Day"]][row["Time Slot"]][fac] = f"<b>{row['Course Code']}</b><br>{row['Batch']}<br><i>{row['Room']}</i>"

                fac_html = "<div style='overflow-x:auto;'><table style='width:100%; border-collapse: collapse; text-align: center; font-size: 13px; font-family: sans-serif;'>"
                fac_html += "<tr><th style='border: 1px solid #ccc; background-color: #1F4E78; color: white; padding: 10px;'>Day</th>"
                fac_html += "<th style='border: 1px solid #ccc; background-color: #1F4E78; color: white; padding: 10px;'>Time Slot</th>"
                for f in selected_facs: fac_html += f"<th style='border: 1px solid #ccc; background-color: #1F4E78; color: white; padding: 10px;'>{f}</th>"
                fac_html += "</tr>"
                for d in days_ordered:
                    for i, s in enumerate(slots_ordered):
                        fac_html += "<tr>"
                        if i == 0: fac_html += f"<td rowspan='4' style='border: 1px solid #ccc; font-weight: bold; background-color: #f8f9fa; vertical-align: middle;'>{d}</td>"
                        fac_html += f"<td style='border: 1px solid #ccc; background-color: #e9ecef; padding: 8px; white-space: nowrap;'>{s}</td>"
                        for f in selected_facs:
                            val = fac_grid[d][s][f]
                            bg = "#e6f2ff" if val else "#ffffff"
                            fac_html += f"<td style='border: 1px solid #ccc; padding: 10px; background-color: {bg}; min-width: 150px;'>{val}</td>"
                        fac_html += "</tr>"
                fac_html += "</table></div>"
                st.markdown(fac_html, unsafe_allow_html=True)

        with tab_swap:
            st.write("### Interactive Manual Swap Engine")
            st.info("Swap time slots between two specific scheduled classes. The system will check for clashes before committing.")
            
            task_list = {f"{r['Day']} {r['Time Slot']} | {r['Batch']} | {r['Course Code']} ({r['Faculty']})": r for r in st.session_state.preview_data}
            
            col_sw1, col_sw2 = st.columns(2)
            with col_sw1:
                swap_1 = st.selectbox("Select First Class to Swap:", options=list(task_list.keys()))
            with col_sw2:
                swap_2 = st.selectbox("Select Second Class to Swap:", options=list(task_list.keys()))
                
            if st.button("🔄 Execute Swap"):
                t1, t2 = task_list[swap_1], task_list[swap_2]
                clash_found = False
                
                for r in st.session_state.preview_data:
                    if r["ID"] not in [t1["ID"], t2["ID"]]:
                        if r["Day"] == t2["Day"] and r["Time Slot"] == t2["Time Slot"]:
                            if r["Faculty"] == t1["Faculty"] and t1["Faculty"] != "TBA": clash_found = True; st.error(f"Clash: {t1['Faculty']} is already teaching at {t2['Day']} {t2['Time Slot']}")
                            if r["Batch"] == t1["Batch"]: clash_found = True; st.error(f"Clash: {t1['Batch']} already has a class at {t2['Day']} {t2['Time Slot']}")
                        if r["Day"] == t1["Day"] and r["Time Slot"] == t1["Time Slot"]:
                            if r["Faculty"] == t2["Faculty"] and t2["Faculty"] != "TBA": clash_found = True; st.error(f"Clash: {t2['Faculty']} is already teaching at {t1['Day']} {t1['Time Slot']}")
                            if r["Batch"] == t2["Batch"]: clash_found = True; st.error(f"Clash: {t2['Batch']} already has a class at {t1['Day']} {t1['Time Slot']}")
                            
                if not clash_found:
                    new_data = copy.deepcopy(st.session_state.preview_data)
                    for r in new_data:
                        if r["ID"] == t1["ID"]:
                            r["Day"], r["Time Slot"] = t2["Day"], t2["Time Slot"]
                        elif r["ID"] == t2["ID"]:
                            r["Day"], r["Time Slot"] = t1["Day"], t1["Time Slot"]
                    st.session_state.routine_history.append(new_data)
                    st.session_state.preview_data = new_data
                    st.success("Swap Successful! The Grid and Faculty views have been updated.")
                    st.rerun()
                    
            if len(st.session_state.routine_history) > 1:
                if st.button("↩️ Undo Last Swap", type="secondary"):
                    st.session_state.routine_history.pop()
                    st.session_state.preview_data = st.session_state.routine_history[-1]
                    st.rerun()
